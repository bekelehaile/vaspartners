#!/usr/bin/env python3
"""
One-shot extractor: finance Excel → JSON snapshots for Laravel data migrations.

Usage (from repo root):
  docker run --rm -v "$PWD:/data" python:3.12-slim bash -c \\
    'pip install -q openpyxl && python /data/scripts/extract_revenue_excel.py'

Re-run only when the Excel changes; migrations read the committed JSON.
"""

from __future__ import annotations

import json
import re
from collections import Counter, OrderedDict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "April 2026 tota;l revenu.xlsx"
OUT_DIR = ROOT / "vaspartners" / "backend" / "database" / "data" / "revenue"

SHEET_CATALOG = {
    "Premium SMS MT": "mt-mobile-terminated-premium",
    "Premium SMS MT ": "mt-mobile-terminated-premium",
    "SMS-MO_Premium": "mo-mobile-originating",
    "Voice_Premium": "voice-premium",
    "Voice": "voice-premium",
    "CRBT-Partners": "crbt",
    "CSA": "api",
    "API with MA": "api",
    "API with MA ": "api",
    "API- Call Signature": "api",
    "API-Mega Promo 130": "api",
}

SHEET_PERIOD_OVERRIDE = {
    "API- Call Signature": "January 2025",
    "API-Mega Promo 130": "January 2025",
    "CRBT-Partners": "March 2026",
}


def norm(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    if not s or s.lower() in {"none", "null", "nan", "-", "n/a"}:
        return None
    return s


def is_total(name: str | None) -> bool:
    return bool(name) and name.strip().upper() == "TOTAL"


def detect_period(rows, sheet: str) -> str | None:
    if sheet in SHEET_PERIOD_OVERRIDE:
        return SHEET_PERIOD_OVERRIDE[sheet]
    blob = " ".join(str(c) for r in rows[:6] for c in r if c is not None)
    m = re.search(
        r"(January|February|March|April|May|June|July|August|September|October|November|December)\s*[\-–]?\s*([0-9]{4})",
        blob,
        re.I,
    )
    if m:
        return f"{m.group(1).title()} {m.group(2)}"
    return None


def parse_amount(v):
    if v is None:
        return None
    try:
        f = float(v)
        if f != f:
            return None
        return round(f, 4)
    except Exception:
        return None


def extract_short_from_name(name: str | None):
    if not name:
        return None, name
    m = re.search(r"\((\d{3,6})\)\s*$", name)
    if not m:
        return None, name
    return m.group(1), name[: m.start()].strip()


def data_start(sheet: str) -> int:
    if sheet.strip() == "Premium SMS MT" or sheet.startswith("Premium SMS MT"):
        return 3
    if "API with MA" in sheet:
        return 4
    if sheet in {"CSA", "API- Call Signature", "API-Mega Promo 130", "SMS-MO_Premium"}:
        return 3
    if sheet in {"Voice_Premium", "CRBT-Partners", "Voice"}:
        return 2
    return 1


def parse_row(sheet: str, r: list):
    service_id = short_code = partner_name = None
    amount = None

    if sheet == "Premium SMS MT":
        short_code, partner_name = norm(r[0]), norm(r[1] if len(r) > 1 else None)
        amount = parse_amount(r[6] if len(r) > 6 else None)
    elif sheet.startswith("Premium SMS MT"):
        service_id = norm(r[0])
        short_code = norm(r[1] if len(r) > 1 else None)
        partner_name = norm(r[2] if len(r) > 2 else None)
        amount = parse_amount(r[14] if len(r) > 14 else None)
        if amount is None:
            amount = parse_amount(r[7] if len(r) > 7 else None)
    elif sheet == "CSA":
        service_id, partner_name = norm(r[0]), norm(r[1] if len(r) > 1 else None)
        amount = parse_amount(r[4] if len(r) > 4 else None)
    elif "API with MA" in sheet:
        service_id, partner_name = norm(r[0]), norm(r[1] if len(r) > 1 else None)
        amount = parse_amount(r[11] if len(r) > 11 else None)
        if amount is None:
            amount = parse_amount(r[5] if len(r) > 5 else None)
    elif sheet == "Voice_Premium":
        short_code, partner_name = norm(r[0]), norm(r[1] if len(r) > 1 else None)
        amount = parse_amount(r[5] if len(r) > 5 else None)
    elif sheet == "API- Call Signature":
        service_id, partner_name = norm(r[0]), norm(r[1] if len(r) > 1 else None)
        amount = parse_amount(r[4] if len(r) > 4 else None)
    elif sheet == "API-Mega Promo 130":
        service_id = norm(r[0])
        partner_name = norm(r[2] if len(r) > 2 else None) or norm(r[1] if len(r) > 1 else None)
        amount = parse_amount(r[5] if len(r) > 5 else None)
    elif sheet == "SMS-MO_Premium":
        short_code, partner_name = norm(r[0]), norm(r[1] if len(r) > 1 else None)
        amount = parse_amount(r[5] if len(r) > 5 else None)
    elif sheet == "CRBT-Partners":
        service_id, partner_name = norm(r[0]), norm(r[1] if len(r) > 1 else None)
        amount = parse_amount(r[5] if len(r) > 5 else None)
    elif sheet == "Voice":
        short_code = norm(r[0])
        amount = parse_amount(r[3] if len(r) > 3 else None)
    else:
        return None

    if is_total(partner_name):
        return None
    if not service_id and not short_code:
        return None

    extracted, cleaned = extract_short_from_name(partner_name)
    if extracted and not short_code:
        short_code = extracted
    if cleaned:
        partner_name = cleaned

    return service_id, short_code, partner_name, amount


def prefer_service_id(current: str | None, incoming: str | None) -> str | None:
    """Prefer the finance-style id (usually longer / zero-padded) when values are equivalent."""
    if not current:
        return incoming
    if not incoming:
        return current
    if current == incoming:
        return current
    cur_core = current.lstrip("0") or "0"
    inc_core = incoming.lstrip("0") or "0"
    if cur_core == inc_core:
        return current if len(current) >= len(incoming) else incoming
    return current


def merge_partner(bucket: OrderedDict, row: dict) -> None:
    sid = row["service_id"]
    sc = row["short_code"]
    effective_sid = sid or sc
    if not effective_sid:
        return

    found = None
    if sid:
        # Match exact or zero-pad-equivalent service id
        core = sid.lstrip("0") or "0"
        for key, p in bucket.items():
            if key == sid or (key.lstrip("0") or "0") == core:
                found = key
                break
    if found is None and sc:
        for key, p in bucket.items():
            if p.get("short_code") == sc:
                found = key
                break
            if not p.get("service_id_from_excel") and p.get("service_id") == sc:
                found = key
                break

    if found is None:
        bucket[effective_sid] = {
            "service_id": effective_sid,
            "service_id_from_excel": bool(sid),
            "short_code": sc,
            "partner_name": row["partner_name"],
            "catalog_slug": row["catalog_slug"] or "api",
            "source_sheets": [row["sheet"]],
        }
        return

    cur = bucket.pop(found)
    chosen_sid = prefer_service_id(cur.get("service_id"), sid if sid else None) or cur["service_id"]
    if sid:
        cur["service_id_from_excel"] = True
    cur["service_id"] = chosen_sid

    if sc and not cur.get("short_code"):
        cur["short_code"] = sc
    if row["partner_name"]:
        if not cur.get("partner_name") or len(row["partner_name"]) >= len(cur.get("partner_name") or ""):
            cur["partner_name"] = row["partner_name"]
    if row["catalog_slug"] and not cur.get("catalog_slug"):
        cur["catalog_slug"] = row["catalog_slug"]
    if row["sheet"] not in cur["source_sheets"]:
        cur["source_sheets"].append(row["sheet"])

    bucket[chosen_sid] = cur


def main() -> None:
    if not XLSX.is_file():
        raise SystemExit(f"Excel not found: {XLSX}")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)

    rows_out: list[dict] = []
    for sheet in wb.sheetnames:
        if sheet not in SHEET_CATALOG:
            print(f"skip unknown sheet: {sheet!r}")
            continue
        ws = wb[sheet]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        period = detect_period(rows, sheet)
        catalog = SHEET_CATALOG[sheet]
        start = data_start(sheet)

        for idx, r in enumerate(rows[start:], start=start + 1):
            if not any(c is not None and str(c).strip() != "" for c in r):
                continue
            parsed = parse_row(sheet, r)
            if not parsed:
                continue
            service_id, short_code, partner_name, amount = parsed
            rows_out.append(
                {
                    "sheet": sheet.strip(),
                    "excel_row": idx,
                    "period": period,
                    "catalog_slug": catalog,
                    "service_id": service_id,
                    "short_code": short_code,
                    "partner_name": partner_name,
                    "amount": amount,
                }
            )

    wb.close()

    bucket: OrderedDict[str, dict] = OrderedDict()
    for row in rows_out:
        merge_partner(bucket, row)

    partners_final = []
    for p in bucket.values():
        partners_final.append(
            {
                "service_id": p["service_id"],
                "short_code": p.get("short_code"),
                "partner_name": p.get("partner_name") or f"Partner {p['service_id']}",
                "catalog_slug": p.get("catalog_slug") or "api",
                "service_id_from_excel": bool(p.get("service_id_from_excel")),
                "source_sheets": p.get("source_sheets", []),
            }
        )

    meta = {
        "source_file": XLSX.name,
        "sheet_count": len({r["sheet"] for r in rows_out}),
        "row_count": len(rows_out),
        "partner_count": len(partners_final),
        "periods": sorted({r["period"] for r in rows_out if r.get("period")}),
    }

    (OUT_DIR / "excel_revenue_rows.json").write_text(
        json.dumps({"meta": meta, "rows": rows_out}, indent=2) + "\n"
    )
    (OUT_DIR / "excel_revenue_partners.json").write_text(
        json.dumps({"meta": meta, "partners": partners_final}, indent=2) + "\n"
    )

    print(json.dumps(meta, indent=2))
    print("partners by catalog:", dict(Counter(p["catalog_slug"] for p in partners_final)))
    print(
        "provisional service_id (short code only):",
        sum(1 for p in partners_final if not p["service_id_from_excel"]),
    )


if __name__ == "__main__":
    main()
